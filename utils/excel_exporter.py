import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHEADER_FILL = PatternFill(start_color="c9a84c", end_color="c9a84c", fill_type="solid")
SUBHEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL = PatternFill(start_color="f0f4ff", end_color="f0f4ff", fill_type="solid")
BORDER = Border(
    left=Side(style='thin', color='d0d8e8'),
    right=Side(style='thin', color='d0d8e8'),
    top=Side(style='thin', color='d0d8e8'),
    bottom=Side(style='thin', color='d0d8e8'),
)

def export_applications_excel(applications, title='Applications_Report'):
    wb = Workbook()
    ws = wb.active
    ws.title = 'OD Applications'

    # Title row
    ws.merge_cells('A1:L1')
    ws['A1'] = 'ADHIYAMAAN COLLEGE OF ENGINEERING - OD Applications Report'
    ws['A1'].font = Font(bold=True, size=14, color='1e3a5f')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:L2')
    ws['A2'] = f'ACOODA - {title.replace("_", " ")}'
    ws['A2'].font = Font(bold=True, size=10, color='888888')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    # Headers
    headers = [
        'App ID', 'Student Name', 'Reg. No.', 'Department',
        'Year', 'Section', 'Event Name', 'Reason', 'From Date',
        'To Date', 'Status', 'Submitted On'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[4].height = 22

    # Data rows
    for row_idx, app in enumerate(applications, 5):
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()
        student = app.student
        row_data = [
            app.application_id,
            student.name if student else '',
            student.register_no if student else '',
            student.department if student else '',
            student.year if student else '',
            student.section if student else '',
            app.event_name,
            app.reason,
            app.from_date.strftime('%d-%m-%Y') if app.from_date else '',
            app.to_date.strftime('%d-%m-%Y') if app.to_date else '',
            app.status,
            app.created_at.strftime('%d-%m-%Y %H:%M') if app.created_at else '',
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            # Color status
            if col == 11:
                if 'Approved' in str(value):
                    cell.font = Font(color='16a34a', bold=True)
                elif 'Rejected' in str(value):
                    cell.font = Font(color='dc2626', bold=True)
                elif 'Pending' in str(value):
                    cell.font = Font(color='d97706', bold=True)
        ws.row_dimensions[row_idx].height = 18

    # Auto-fit columns
    col_widths = [16, 22, 14, 20, 8, 10, 28, 18, 14, 14, 22, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header
    ws.freeze_panes = 'A5'

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', prefix='ACOODA_')
    tmp.close()
    wb.save(tmp.name)
    return tmp.name
